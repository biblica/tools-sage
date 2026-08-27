"""Immutable import adapters for governed SAGE-local project-index sources."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from ..atomic import atomic_write_json
from ..errors import ValidationError
from ..hashing import sha256_bytes
from ..registry import EcosystemConfig
from .authority_registry import authority_source_for_type
from .policy import IMPORT_LIFT_STATUS
from .semdom import normalise_authority_json, normalise_specific_first_docx, split_semdom
from .store import authority_root, ensure_authority_selected, ensure_import_active, import_root, safe_id, snapshot_file, write_manifest


def _stable_id(*parts: str) -> str:
    """Return a compact deterministic identifier for imported semantic records."""
    payload = "\x1f".join(parts).encode("utf-8")
    return sha256_bytes(payload)[:24]


def _bool_yes(value: object) -> bool:
    """Interpret common affirmative spreadsheet values deterministically."""
    return str(value or "").strip().casefold() in {"yes", "y", "true", "1"}


def _write_import(
    config: EcosystemConfig,
    *,
    language: str,
    source_id: str,
    source_type: str,
    source_path: Path,
    records: list[dict[str, Any]],
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Persist one immutable normalized import with source provenance and hash."""
    root = import_root(config, language, source_id)
    root.mkdir(parents=True, exist_ok=True)
    suffix = source_path.suffix.lower() or ".dat"
    provenance = snapshot_file(source_path, root / f"source{suffix}")
    manifest_path = root / "manifest.json"
    records_path = root / "records.json"
    if manifest_path.is_file() or records_path.is_file():
        if not manifest_path.is_file() or not records_path.is_file():
            raise ValidationError(f"Incomplete immutable semantic import snapshot: {root}")
        try:
            existing = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError) as exc:
            raise ValidationError(f"Invalid immutable semantic import manifest {manifest_path}: {exc}") from exc
        if not isinstance(existing, dict) or existing.get("source", {}).get("sha256") != provenance["sha256"]:
            raise ValidationError(f"Immutable semantic import ID already exists with different content: {source_id}")
        if str(existing.get("source_type", "")) != source_type or str(existing.get("language", "")) != language:
            raise ValidationError(f"Immutable semantic import ID is already bound to another type or language: {source_id}")
        ensure_import_active(config, language=language, source_id=source_id)
        return {**existing, "import_root": str(root), "status_detail": "UNCHANGED_IMMUTABLE_SNAPSHOT"}
    atomic_write_json(records_path, records)
    manifest = {
        "schema_version": "1.0",
        "source_id": source_id,
        "source_type": source_type,
        "language": language,
        "authority": "PROJECT_INDEX_ONLY",
        "evidence_class": "PROJECT_INDEX_EVIDENCE",
        "translation_authority": False,
        "scripture_authority": False,
        "mutable": False,
        "record_count": len(records),
        "records_sha256": sha256_bytes((json.dumps(records, ensure_ascii=False, sort_keys=True) + "\n").encode("utf-8")),
        "source": provenance,
        **(extra or {}),
    }
    write_manifest(manifest_path, manifest)
    ensure_import_active(config, language=language, source_id=source_id)
    return {**manifest, "import_root": str(root), "status_detail": "IMPORTED_IMMUTABLE_SNAPSHOT"}


def import_rwc_seed_xlsx(
    config: EcosystemConfig,
    path: Path,
    *,
    source_id: str,
    language: str = "KKH",
    analysis_language: str = "en",
    sheet: str = "Entire Luke",
    headword_column: str = "Text",
    gloss_column: str = "English gloss",
    key_term_column: str = "Key Term",
    semdom_column: str = "SIL SemDom",
) -> dict[str, Any]:
    """Import one operator-declared RWC/SemDom workbook as non-authoritative seed evidence."""
    path = path.resolve()
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise ValidationError(f"RWC seed workbook does not contain sheet {sheet!r}")
    worksheet = workbook[sheet]
    header = [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    column_names = {
        "headword": headword_column,
        "gloss": gloss_column,
        "key_term": key_term_column,
        "semdom": semdom_column,
    }
    missing = [name for name in column_names.values() if name not in header]
    if missing:
        raise ValidationError(f"RWC seed workbook is missing configured columns: {missing!r}; found {header!r}")
    columns = {key: header.index(name) for key, name in column_names.items()}
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        headword = str(row[columns["headword"]] or "").strip()
        if not headword:
            continue
        gloss = str(row[columns["gloss"]] or "").strip()
        code, label = split_semdom(row[columns["semdom"]])
        sense_id = _stable_id(source_id, headword, code, gloss)
        records.append(
            {
                "record_id": _stable_id(source_id, str(row_number), headword, code),
                "record_type": "RWC_SEED_SENSE",
                "language": language,
                "headword": headword,
                "lemma": headword,
                "lemma_status": "SEED_HEADWORD",
                "surface_forms": [headword],
                "key_term": _bool_yes(row[columns["key_term"]]),
                "senses": [
                    {
                        "sense_id": sense_id,
                        "glosses": {analysis_language: gloss} if gloss else {},
                        "definitions": {},
                        "semdom": [{"code": code, "label": label, "role": "PRIMARY"}],
                        "part_of_speech": None,
                        "references": [],
                        "status": "SEED",
                    }
                ],
                "provenance": {
                    "source_id": source_id,
                    "source_type": "RWC_SEMDOM_XLSX",
                    "row": row_number,
                },
            }
        )
    return _write_import(
        config,
        language=language,
        source_id=safe_id(source_id, "source_id"),
        source_type="RWC_SEMDOM_XLSX",
        source_path=path,
        records=records,
        extra={
            "analysis_language": analysis_language,
            "sheet": sheet,
            "status": "SEED",
            "columns": column_names,
        },
    )


def _local_name(element: ET.Element) -> str:
    """Return one XML local tag name so LIFT namespaces do not affect interchange."""
    return element.tag.rsplit("}", 1)[-1]


def _children(parent: ET.Element | None, name: str) -> list[ET.Element]:
    """Return direct children matching one local LIFT element name."""
    if parent is None:
        return []
    return [child for child in list(parent) if _local_name(child) == name]


def _child(parent: ET.Element | None, name: str) -> ET.Element | None:
    """Return the first direct child matching one local LIFT element name."""
    children = _children(parent, name)
    return children[0] if children else None


def _child_text(parent: ET.Element, tag: str) -> str:
    """Return normalized text from one direct LIFT child."""
    node = _child(parent, tag)
    return str(node.text or "").strip() if node is not None else ""


def _forms(container: ET.Element | None) -> dict[str, str]:
    """Return language-to-text values from one LIFT multi-form container."""
    result: dict[str, str] = {}
    if container is None:
        return result
    for form in _children(container, "form"):
        lang = str(form.attrib.get("lang", "")).strip()
        text = _child_text(form, "text")
        if lang and text:
            result[lang] = text
    return result

def import_lift_snapshot(
    config: EcosystemConfig,
    path: Path,
    *,
    source_id: str,
    source_application: str,
    language: str,
) -> dict[str, Any]:
    """Import a FLEx or The Combine LIFT file into the richer SAGE semantic model."""
    path = path.resolve()
    application = str(source_application).strip().upper()
    if application not in {"FLEX", "COMBINE"}:
        raise ValidationError("source_application must be FLEx or Combine")
    try:
        tree = ET.parse(path)
    except (OSError, ET.ParseError) as exc:
        raise ValidationError(f"Invalid LIFT file {path}: {exc}") from exc
    root = tree.getroot()
    if root.tag.rsplit("}", 1)[-1] != "lift":
        raise ValidationError("LIFT root element must be <lift>")
    records: list[dict[str, Any]] = []
    for entry_index, entry in enumerate(_children(root, "entry"), start=1):
        lexical = _forms(_child(entry, "lexical-unit"))
        citation = _forms(_child(entry, "citation"))
        headword = lexical.get(language) or citation.get(language)
        if not headword:
            continue
        source_entry_id = str(entry.attrib.get("id") or entry.attrib.get("guid") or _stable_id(headword, str(entry_index)))
        entry_id = _stable_id(source_id, source_entry_id)
        senses: list[dict[str, Any]] = []
        for sense_index, sense in enumerate(_children(entry, "sense"), start=1):
            semdom: list[dict[str, str]] = []
            for trait in _children(sense, "trait"):
                name = str(trait.attrib.get("name", "")).strip()
                if name not in {"semantic-domain-ddp4", "semantic-domain"}:
                    continue
                value = str(trait.attrib.get("value", "")).strip()
                try:
                    code, label = split_semdom(value)
                except ValidationError:
                    code, label = value, ""
                semdom.append({"code": code, "label": label, "role": "IMPORTED"})
            glosses = _forms(_child(sense, "gloss"))
            definitions = _forms(_child(sense, "definition"))
            grammatical = _child(sense, "grammatical-info")
            part_of_speech = (
                str(grammatical.attrib.get("value", "")).strip()
                if grammatical is not None
                else None
            )
            notes: list[str] = []
            for note in _children(sense, "note"):
                note_forms = _forms(note)
                notes.extend(value for value in note_forms.values() if value)
            source_sense_id = str(sense.attrib.get("id") or _stable_id(source_entry_id, str(sense_index)))
            senses.append(
                {
                    "sense_id": _stable_id(source_id, source_sense_id),
                    "source_sense_id": source_sense_id,
                    "glosses": glosses,
                    "definitions": definitions,
                    "semdom": semdom,
                    "part_of_speech": part_of_speech,
                    "references": [],
                    "notes": notes,
                    "status": IMPORT_LIFT_STATUS,
                }
            )
        records.append(
            {
                "record_id": entry_id,
                "record_type": f"{application}_LIFT_ENTRY",
                "language": language,
                "headword": headword,
                "lemma": headword,
                "lemma_status": "IMPORTED_LEXEME",
                "surface_forms": sorted(set(lexical.values()) | set(citation.values())),
                "key_term": False,
                "senses": senses or [
                    {
                        "sense_id": _stable_id(source_id, entry_id, "sense"),
                        "glosses": {},
                        "definitions": {},
                        "semdom": [],
                        "part_of_speech": None,
                        "references": [],
                        "notes": [],
                        "status": IMPORT_LIFT_STATUS,
                    }
                ],
                "provenance": {
                    "source_id": source_id,
                    "source_type": f"{application}_LIFT",
                    "source_entry_id": source_entry_id,
                },
            }
        )
    return _write_import(
        config,
        language=language,
        source_id=safe_id(source_id, "source_id"),
        source_type=f"{application}_LIFT",
        source_path=path,
        records=records,
        extra={"source_application": "FLEx" if application == "FLEX" else "Combine", "status": IMPORT_LIFT_STATUS},
    )


def import_semdom_authority_json(
    config: EcosystemConfig,
    path: Path,
    *,
    source_id: str | None = None,
) -> dict[str, Any]:
    """Import an operator-supplied current SIL semdom.org JSON snapshot as classification authority."""
    spec = authority_source_for_type(config, "semdom")
    resolved_source_id = safe_id(source_id or spec.default_source_id, "source_id")
    path = path.resolve()
    records = normalise_authority_json(path)
    root = authority_root(config) / spec.storage_directory / resolved_source_id
    root.mkdir(parents=True, exist_ok=True)
    provenance = snapshot_file(path, root / "source.json")
    atomic_write_json(root / "domains.json", records)
    manifest = {
        "schema_version": "1.0",
        "source_id": resolved_source_id,
        "source_type": "SIL_SEMDOM_AUTHORITY",
        "semantic_authority": True,
        "translation_authority": False,
        "domain_count": len(records),
        "source": provenance,
    }
    write_manifest(root / "manifest.json", manifest)
    ensure_authority_selected(config, authority_type="semdom", source_id=resolved_source_id)
    return {**manifest, "authority_root": str(root)}


def import_specific_first_docx(
    config: EcosystemConfig,
    path: Path,
    *,
    source_id: str | None = None,
) -> dict[str, Any]:
    """Import RapidWords specific-first folder divisions as retrieval metadata only."""
    spec = authority_source_for_type(config, "folders")
    resolved_source_id = safe_id(source_id or spec.default_source_id, "source_id")
    path = path.resolve()
    folders = normalise_specific_first_docx(path)
    root = authority_root(config) / spec.storage_directory / resolved_source_id
    root.mkdir(parents=True, exist_ok=True)
    provenance = snapshot_file(path, root / "source.docx")
    atomic_write_json(root / "folders.json", folders)
    manifest = {
        "schema_version": "1.0",
        "source_id": resolved_source_id,
        "source_type": "RAPIDWORDS_SPECIFIC_FIRST_FOLDER_MAP",
        "semantic_authority": False,
        "translation_authority": False,
        "processing_metadata_only": True,
        "folder_count": len(folders),
        "domain_count": sum(len(folder["domains"]) for folder in folders),
        "source": provenance,
    }
    write_manifest(root / "manifest.json", manifest)
    ensure_authority_selected(config, authority_type="folders", source_id=resolved_source_id)
    return {**manifest, "authority_root": str(root)}
